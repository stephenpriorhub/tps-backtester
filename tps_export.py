"""
tps_export.py
=============
Export TPS backtest results to Excel matching the TradingView v3 "list of
trades" layout: ONE entry row + ONE exit row per signal, with the exit row
carrying the per-trade Total Return / Max Return / Duration and the exit label
(Long / RSI Exit / Stop / Open).

Public API
----------
    export_to_excel(all_trades, cfg, output_path) -> str
    compute_signal_stats(df) -> dict          # per-ticker summary metrics
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Reference ticker order (Summary sheet columns) ────────────────────────────
REFERENCE_TICKERS = ["AAPL", "GOOGL", "NVDA", "TSLA", "AMZN", "MSFT", "AMD", "ORCL", "NFLX"]

# ── Number formats ─────────────────────────────────────────────────────────────
FMT_PCT       = "0.00%"
FMT_PCT3      = "0.000%"
FMT_DECIMAL   = "0.00"
FMT_DATETIME  = r'yyyy\-mm\-dd\ hh:mm'
FMT_NUMBER    = "#,##0.00"

# ── Styling ──────────────────────────────────────────────────────────────────
FONT_NORMAL = Font(name="Calibri", size=11)
FONT_BOLD   = Font(name="Calibri", size=11, bold=True)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


# ══════════════════════════════════════════════════════════════════════════════
# PER-SIGNAL STATS  (single source of truth for both workbook + on-screen summary)
# ══════════════════════════════════════════════════════════════════════════════

def _signal_frame(df: pd.DataFrame) -> pd.DataFrame:
    """Return just the Exit rows — one per signal in v3."""
    if df is None or df.empty:
        return pd.DataFrame()
    return df[df["Type"] == "Exit long"].copy()


def compute_signal_stats(df: pd.DataFrame) -> dict:
    """
    Per-ticker summary metrics (per-signal), computed from the Exit rows.
    Matches the workbook summary tab and the on-screen dashboard.
    """
    keys = ("count", "wins", "win_rate", "avg_ret", "avg_win", "avg_loss",
            "avg_max_ret", "win_rate_max", "expectancy", "avg_dur")
    exits = _signal_frame(df)
    if exits.empty:
        return {k: 0 for k in keys}

    ret      = exits["Net P&L %"] / 100.0                # fraction
    max_ret  = exits["Max Return"].fillna(0.0)           # fraction
    n        = len(exits)
    wins     = int((ret > 0).sum())
    win_rate = wins / n if n else 0.0
    pos      = ret[ret > 0]
    neg      = ret[ret < 0]
    avg_win  = pos.mean() if len(pos) else 0.0
    avg_loss = neg.mean() if len(neg) else 0.0

    return {
        "count":        n,
        "wins":         wins,
        "win_rate":     win_rate,
        "avg_ret":      ret.mean(),
        "avg_win":      avg_win,
        "avg_loss":     avg_loss,
        "avg_max_ret":  max_ret.mean() if n else 0.0,
        "win_rate_max": (max_ret > 0).sum() / n if n else 0.0,
        "expectancy":   win_rate * avg_win + (1 - win_rate) * avg_loss,
        "avg_dur":      exits["Duration"].mean() if n else 0.0,
    }


# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY SHEET
# ══════════════════════════════════════════════════════════════════════════════

def _set_cell(ws, row, col, value, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    if fmt:
        cell.number_format = fmt
    cell.font = FONT_BOLD if bold else FONT_NORMAL
    return cell


def _auto_width(ws, min_width=8, max_width=30, overrides=None):
    overrides = overrides or {}
    for col in ws.columns:
        letter = col[0].column_letter
        if letter in overrides:
            ws.column_dimensions[letter].width = overrides[letter]
            continue
        max_len = max((len(str(c.value or "")) for c in col), default=min_width)
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_width), max_width)


def _write_summary_sheet(wb, tickers, chart_tf, stats_by_ticker):
    ws = wb.create_sheet(f"{chart_tf}min Summary", 0)
    n = len(tickers)
    avgs_col = n + 2

    for i, ticker in enumerate(tickers):
        _set_cell(ws, 1, i + 2, ticker, bold=True)
    _set_cell(ws, 1, avgs_col, f"V3 ({chart_tf})avgs", bold=True)

    metric_rows = [
        ("Count",              "count",        FMT_NUMBER, "sum"),
        ("# Wins",             "wins",         FMT_NUMBER, "sum"),
        ("Win Rate",           "win_rate",     FMT_PCT,    "avg"),
        ("Avg Return %",       "avg_ret",      FMT_PCT,    "avg"),
        ("Avg Win %",          "avg_win",      FMT_PCT,    "avg"),
        ("Avg Loss %",         "avg_loss",     FMT_PCT,    "avg"),
        ("Avg Max Return %",   "avg_max_ret",  FMT_PCT,    "avg"),
        ("Win Rate (max)",     "win_rate_max", FMT_PCT,    "avg"),
        ("Expectancy:",        "expectancy",   FMT_PCT,    "avg"),
        ("Avg Duration (days)","avg_dur",      FMT_NUMBER, "avg"),
    ]
    for offset, (label, *_ ) in enumerate(metric_rows):
        _set_cell(ws, offset + 2, 1, label, bold=True)

    for i, ticker in enumerate(tickers):
        col = i + 2
        stats = stats_by_ticker.get(ticker) or {}
        for r, (_, key, fmt, _agg) in enumerate(metric_rows, start=2):
            _set_cell(ws, r, col, stats.get(key, 0), fmt=fmt)

    for r, (_, key, fmt, agg) in enumerate(metric_rows, start=2):
        vals = [(stats_by_ticker.get(t) or {}).get(key, 0) for t in tickers]
        out = sum(vals) if agg == "sum" else (sum(vals) / len(vals) if vals else 0)
        _set_cell(ws, r, avgs_col, out, fmt=fmt)

    _auto_width(ws, overrides={"A": 20})


# ══════════════════════════════════════════════════════════════════════════════
# TICKER SHEET  (dump the engine's 2-row-per-trade DataFrame)
# ══════════════════════════════════════════════════════════════════════════════

TICKER_COLUMNS = [
    "Trade #", "Type", "Date and time", "Signal", "Price USD",
    "Size (qty)", "Size (value)", "Total Return", "Max Return", "Duration",
    "Net P&L USD", "Net P&L %",
    "Favorable excursion USD", "Favorable excursion %",
    "Adverse excursion USD", "Adverse excursion %",
    "Cumulative P&L USD", "Cumulative P&L %",
]

WIN_FILL  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


def _write_ticker_sheet(wb, ticker, df):
    ws = wb.create_sheet(ticker)

    for c_idx, header in enumerate(TICKER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    if df is None or df.empty:
        ws.freeze_panes = "A2"
        return compute_signal_stats(df)

    pct_cols = {"Total Return", "Max Return"}
    dec_cols = {"Net P&L %", "Favorable excursion %", "Adverse excursion %",
                "Cumulative P&L %", "Duration"}

    r_idx = 2
    for _, row in df.iterrows():
        is_exit = row["Type"] == "Exit long"
        for c_idx, col in enumerate(TICKER_COLUMNS, start=1):
            val = row.get(col)
            if isinstance(val, float) and pd.isna(val):
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = FONT_NORMAL
            if col in pct_cols and val is not None:
                cell.number_format = FMT_PCT
            elif col in dec_cols and val is not None:
                cell.number_format = FMT_DECIMAL
        # tint exit rows by win/loss
        if is_exit:
            pnl = row.get("Net P&L %") or 0
            fill = WIN_FILL if pnl > 0 else LOSS_FILL
            for c_idx in range(1, len(TICKER_COLUMNS) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = fill
        r_idx += 1

    ws.freeze_panes = "A2"
    _auto_width(ws, min_width=6, overrides={"C": 18, "M": 22, "O": 20})
    return compute_signal_stats(df)


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════

def export_to_excel(all_trades: dict, cfg: dict, output_path: str) -> str:
    """
    Export TPS v3 backtest results to Excel.

    Parameters
    ----------
    all_trades : {ticker: DataFrame from tps_engine.run_ticker()[1]}
    cfg        : config dict used for the run (chart_tf, score_threshold, ...)
    output_path: destination .xlsx path (created/overwritten)
    """
    chart_tf = int(cfg.get("chart_tf", 78))

    available = set(all_trades.keys())
    ordered = [t for t in REFERENCE_TICKERS if t in available]
    ordered.extend(sorted(t for t in all_trades if t not in set(REFERENCE_TICKERS)))

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    stats_by_ticker = {}
    for ticker in ordered:
        stats_by_ticker[ticker] = _write_ticker_sheet(wb, ticker, all_trades.get(ticker))

    _write_summary_sheet(wb, ordered, chart_tf, stats_by_ticker)

    out = Path(output_path).resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)
