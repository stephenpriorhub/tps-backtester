"""
TPS Long Strategy v3 — Backtester powered by Massive Market Data
=================================================================
Fetches 1-MINUTE OHLCV data from Massive, runs the TPS v3 backtest engine
(RSI-momentum exit), and exports results to an Excel spreadsheet matching the
TradingView list-of-trades format.

QUICK START:
  1. Get your Massive API key from https://massive.com/dashboard
  2. export MASSIVE_API_KEY=your_key_here
  3. python3 tps_run_massive.py

DATA CACHING:
  Downloaded bars are cached as Parquet files in ./data/.
  Subsequent runs (or improvement-idea runs) load from cache in ~1s/ticker.
  Delete the cache folder to force a fresh download.

IMPROVEMENT IDEAS:
  Uncomment one CONFIG block near the top to test an idea in isolation.

UNIVERSE:
  Default = full Nasdaq-100 (101 securities, both GOOG+GOOGL).
  Override by setting CONFIG["tickers"] to any list you like.
"""

import os
import sys
import time
import hashlib
import requests
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from pathlib import Path
from datetime import datetime
from urllib.parse import urlparse, parse_qs


# ──────────────────────────────────────────────────────────────────────────────
# Bootstrap: load API key from .env files so standalone `python3 tps_run_massive.py`
# works without requiring `export MASSIVE_API_KEY=...` in the shell first.
# tps_app.py (Streamlit) patches MASSIVE_API_KEY separately via _load_api_key();
# this bootstrap only sets os.environ if neither key is already present.
# ──────────────────────────────────────────────────────────────────────────────
def _bootstrap_env():
    candidates = [
        Path(__file__).parent / ".env",
        Path.home() / "Documents/GitHub/trading-scanner/.env",
        Path.home() / "Documents/github/trading-scanner/.env",
    ]
    for p in candidates:
        if p.exists():
            for line in p.read_text().splitlines():
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip().strip('"').strip("'")
                if k in ("MASSIVE_API_KEY", "POLYGON_API_KEY") and not os.environ.get(k):
                    os.environ[k] = v
            break  # stop after first found .env


_bootstrap_env()

from tps_engine import (
    DEFAULT_CONFIG, filter_market_hours, resample_ohlcv,
    run_ticker, compute_summary, trades_to_dataframe,
    add_tps_signals,
)

# ──────────────────────────────────────────────────────────────────────────────
# NASDAQ-100 UNIVERSE  (~101 securities — GOOG + GOOGL both included)
# Composition as of early 2025; NDX rebalances quarterly so this is approximate.
# Add, remove, or swap tickers freely — this is just the default starting list.
# ──────────────────────────────────────────────────────────────────────────────

NDX100 = [
    # Mega-cap tech
    "AAPL", "MSFT", "NVDA", "AMZN", "GOOG", "GOOGL", "META", "TSLA", "AVGO",
    # Consumer / discretionary
    "COST", "NFLX", "BKNG", "ORLY", "SBUX", "ROST", "MAR", "PCAR", "ABNB",
    "LULU", "FAST", "DLTR", "CPRT", "TSCO", "EXPE",
    # Semiconductors
    "AMD", "QCOM", "TXN", "AMAT", "ADI", "MU", "LRCX", "KLAC", "NXPI",
    "MRVL", "MCHP", "ON", "INTC", "ASML", "ARM", "GFS", "SMCI",
    # Software / SaaS / Cybersecurity
    "ADBE", "INTU", "SNPS", "CDNS", "PANW", "FTNT", "ZS", "DDOG", "CRWD",
    "TEAM", "WDAY", "TTD", "CSGP", "ANSS", "PLTR", "ORCL",
    # Communication
    "TMUS", "CHTR", "WBD",
    # Healthcare / biotech
    "AMGN", "ISRG", "VRTX", "GILD", "REGN", "IDXX", "DXCM", "BIIB",
    "ILMN", "MRNA", "AZN", "GEHC", "ALGN",
    # Consumer staples
    "MDLZ", "KDP", "MNST", "PEP", "CCEP",
    # Financials / payments
    "PYPL", "CTAS", "PAYX",
    # Industrials / transport
    "CSX", "ODFL", "VRSK", "PCAR",
    # Energy / utilities
    "FANG", "BKR", "EXC", "CEG",
    # Other / diversified
    "TTWO", "CTSH", "MELI", "DASH", "EA", "NDAQ",
]

# De-duplicate (PCAR appears in two groups above)
NDX100 = list(dict.fromkeys(NDX100))

# ──────────────────────────────────────────────────────────────────────────────
# ★ CONFIG — edit this block to change what the backtest runs
# ──────────────────────────────────────────────────────────────────────────────

MASSIVE_API_KEY = os.environ.get("MASSIVE_API_KEY") or os.environ.get("POLYGON_API_KEY", "")
MASSIVE_BASE    = "https://api.massive.com"

CONFIG = {
    **DEFAULT_CONFIG,

    # ── Universe: NDX100 by default, or set to any list
    "tickers": NDX100,

    # ── Date range
    # Polygon.io free tier: data available from ~2016 for daily, ~2021 for 15m intraday.
    # Paid Starter/Developer plans extend intraday back to 2003.
    # Default: 10 years back. Set "start_date" in CLI / app to override.
    "start_date": (datetime.today().replace(year=datetime.today().year - 10)).strftime("%Y-%m-%d"),
    "end_date":   datetime.today().strftime("%Y-%m-%d"),

    # ── Chart timeframe: 78 or 195 (minutes)
    "chart_tf": 78,

    # ── Output file
    "output_file": "TPS Backtest Results NDX100.xlsx",

    # ── Excel detail sheets: how many per-ticker trade logs to write.
    #    They are sorted by win_rate × avg_return (best first).
    #    Set to -1 for all tickers, 0 for summary-only (fastest).
    "detail_sheets": 25,
}

# ── To test an improvement idea, uncomment ONE block:

# IDEA 1 — Looser RSI exit (let winners run further)
# CONFIG["rsi_exit_threshold"] = 50.0

# IDEA 2 — Score threshold raised to 70
# CONFIG["score_threshold"] = 70.0

# IDEA 3 — Squeeze weights rebalanced toward chart TF
# CONFIG["sqz_pts_78"] = 20.0
# CONFIG["sqz_pts_30"] = 20.0
# CONFIG["sqz_pts_15"] = 10.0

# IDEA 4 — Volume filter (volume > 20-bar SMA)
# CONFIG["volume_filter"] = True

# IDEA 5 — Squeeze duration sweet-spot
# CONFIG["min_sqz_bars"] = 5
# CONFIG["max_sqz_bars"] = 12

DATA_DIR = Path(__file__).parent / "data"

# ──────────────────────────────────────────────────────────────────────────────
# MASSIVE API — data fetching with auto-pagination, retry, and disk cache
# ──────────────────────────────────────────────────────────────────────────────

MAX_RETRIES     = 3          # retries per page on transient errors
PAGE_SLEEP      = 0.15       # seconds between paginated requests
TICKER_SLEEP    = 0.5        # seconds between tickers (rate-limit courtesy)


def _cache_path(ticker: str, multiplier: int, timespan: str,
                start: str, end: str) -> Path:
    key = f"{ticker}_{multiplier}_{timespan}_{start}_{end}"
    h = hashlib.md5(key.encode()).hexdigest()[:8]
    return DATA_DIR / f"{ticker}_{multiplier}{timespan}_{start[:7]}_{end[:7]}_{h}.parquet"


def fetch_bars(ticker: str, multiplier: int, timespan: str,
               start: str, end: str, use_cache: bool = True) -> pd.DataFrame:
    """
    Fetch OHLCV bars from Massive with auto-pagination, retry, and disk cache.
    Returns DataFrame with UTC DatetimeIndex and columns: o, h, l, c, v.
    """
    DATA_DIR.mkdir(exist_ok=True)
    cache = _cache_path(ticker, multiplier, timespan, start, end)

    if use_cache and cache.exists():
        df = pd.read_parquet(cache)
        if df.index.tz is None:
            df.index = df.index.tz_localize("UTC")
        return df

    if not MASSIVE_API_KEY:
        # Raise instead of sys.exit — sys.exit inside the Streamlit app would
        # kill the whole server process, not just this fetch.
        raise RuntimeError(
            "MASSIVE_API_KEY is not set. Get your key at "
            "https://massive.com/dashboard and export MASSIVE_API_KEY=..."
        )

    print(f"    fetching {ticker} {multiplier}{timespan}...", end="", flush=True)

    all_bars = []
    path = f"/v2/aggs/ticker/{ticker}/range/{multiplier}/{timespan}/{start}/{end}"
    params = {
        "apiKey":   MASSIVE_API_KEY,
        "adjusted": "true",
        "sort":     "asc",
        "limit":    50000,
    }

    while True:
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                resp = requests.get(MASSIVE_BASE + path, params=params, timeout=30)
                if resp.status_code == 429:
                    wait = 2 ** attempt
                    print(f"[rate-limit, waiting {wait}s]", end="", flush=True)
                    time.sleep(wait)
                    continue
                resp.raise_for_status()
                break
            except requests.RequestException as e:
                if attempt == MAX_RETRIES:
                    print(f" ERROR: {e}")
                    return pd.DataFrame(columns=["o", "h", "l", "c", "v"])
                time.sleep(2 ** attempt)
        else:
            break

        data    = resp.json()
        results = data.get("results") or []
        all_bars.extend(results)
        print(".", end="", flush=True)

        next_url = data.get("next_url")
        if not next_url or not results:
            break

        parsed = urlparse(next_url)
        cursor = parse_qs(parsed.query).get("cursor", [None])[0]
        if not cursor:
            break
        params = {"apiKey": MASSIVE_API_KEY, "cursor": cursor, "limit": 50000}
        path   = parsed.path
        time.sleep(PAGE_SLEEP)

    print(f" {len(all_bars):,} bars")

    if not all_bars:
        return pd.DataFrame(columns=["o", "h", "l", "c", "v"])

    df = pd.DataFrame(all_bars)
    df["t"] = pd.to_datetime(df["t"], unit="ms", utc=True)
    df = df.set_index("t").sort_index()
    df = df[["o", "h", "l", "c", "v"]].astype(float)
    df.to_parquet(cache)
    return df


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL     = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
WIN_FILL     = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LOSS_FILL    = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
TOP_FILL     = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _write_sheet(wb: openpyxl.Workbook, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(sheet_name)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
            elif r_idx % 2 == 0:
                cell.fill = ALT_FILL

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)

    return ws


def _write_ranked_summary(wb: openpyxl.Workbook,
                          ticker_stats: list[dict],
                          chart_tf: int) -> None:
    """Write a full ranked summary sheet: one row per ticker, sorted by score."""
    ws = wb.create_sheet(f"{chart_tf}min NDX Ranking", 0)

    headers = [
        "Rank", "Ticker", "Trades", "Win Rate", "Avg Return %",
        "Avg Win %", "Avg Loss %", "Expectancy", "Avg Duration (days)",
        "Best Trade %", "Worst Trade %",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for rank, row in enumerate(ticker_stats, start=1):
        ws.append([
            rank,
            row["ticker"],
            row["n"],
            row["wr"],
            row["avg_r"],
            row["avg_win"],
            row["avg_loss"],
            row["exp"],
            row["avg_dur"],
            row["best"],
            row["worst"],
        ])
        ws_row = ws[ws.max_row]
        fill = TOP_FILL if rank <= 10 else (ALT_FILL if rank % 2 == 0 else None)
        if fill:
            for cell in ws_row:
                cell.fill = fill

    # Number formats
    pct_cols  = [4, 5, 6, 7, 10, 11]   # columns that are percentages
    for col_idx in pct_cols:
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = "0.000%"
    # Win rate as 0.0%
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = "0.0%"

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 18)


def _compute_ticker_stats(ticker: str, df_trades: pd.DataFrame) -> dict:
    """Compute per-ticker summary stats from the trades DataFrame."""
    empty = dict(ticker=ticker, n=0, wr=0, avg_r=0, avg_win=0,
                 avg_loss=0, exp=0, avg_dur=0, best=0, worst=0, score=0)
    if df_trades is None or df_trades.empty:
        return empty

    exits = df_trades[df_trades["Type"] == "Exit long"]
    if exits.empty:
        return empty

    pnl  = exits["Net P&L %"]
    wins = pnl[pnl > 0]
    loss = pnl[pnl <= 0]
    n    = len(exits)
    wr   = len(wins) / n if n else 0
    avg_r   = pnl.mean()
    avg_win = wins.mean() if len(wins) else 0
    avg_loss= loss.mean() if len(loss) else 0
    exp  = wr * avg_win + (1 - wr) * abs(avg_loss) * -1 if n else 0
    avg_dur = exits["Duration"].mean() if "Duration" in exits.columns else 0
    score   = wr * avg_r   # sort key

    return dict(
        ticker=ticker, n=n, wr=wr, avg_r=avg_r, avg_win=avg_win,
        avg_loss=avg_loss, exp=exp, avg_dur=avg_dur,
        best=pnl.max(), worst=pnl.min(), score=score,
    )


def export_results(all_trades: dict, output_path: str, chart_tf: int,
                   detail_sheets: int = 25) -> None:
    """
    Export results to Excel.
    - Sheet 0: Full Nasdaq-100 ranking (all tickers, sorted best→worst)
    - Sheets 1…N: Per-ticker trade logs for the top `detail_sheets` tickers
      (sorted by win_rate × avg_return). Set detail_sheets=0 for summary only,
      -1 for all tickers.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Build stats for every ticker
    ticker_stats = []
    for ticker, df_trades in all_trades.items():
        ticker_stats.append(_compute_ticker_stats(ticker, df_trades))

    # Sort best → worst by score (WR × avg_return)
    ticker_stats.sort(key=lambda r: r["score"], reverse=True)

    # Ranked summary (always written)
    _write_ranked_summary(wb, ticker_stats, chart_tf)

    # Per-ticker detail sheets
    n_detail = len(ticker_stats) if detail_sheets == -1 else detail_sheets
    written  = 0
    for row in ticker_stats:
        if written >= n_detail:
            break
        ticker    = row["ticker"]
        df_trades = all_trades.get(ticker)
        if df_trades is None or df_trades.empty:
            continue
        _write_sheet(wb, ticker, df_trades)
        written += 1

    wb.save(output_path)
    print(f"\n✓ Saved → {output_path}")
    print(f"  ({len(ticker_stats)} tickers in ranking, {written} detail sheets)")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main():
    tickers   = CONFIG["tickers"]
    start     = CONFIG["start_date"]
    end       = CONFIG["end_date"]
    chart_tf  = CONFIG["chart_tf"]
    n_total   = len(tickers)

    print("=" * 65)
    print(f"  TPS Long Strategy v3 — Massive Market Data Backtest")
    print(f"  Universe : {n_total} tickers")
    print(f"  Period   : {start} → {end}")
    print(f"  Chart TF : {chart_tf}m    Score ≥ {CONFIG['score_threshold']}")
    print(f"  Output   : {CONFIG['output_file']}")
    print("=" * 65)

    all_trades   = {}
    times        = []
    total_t0     = time.time()
    skipped      = []

    for i, ticker in enumerate(tickers, start=1):
        t0 = time.time()
        eta_str = ""
        if times:
            avg_t  = sum(times) / len(times)
            remain = avg_t * (n_total - i + 1)
            m, s   = divmod(int(remain), 60)
            eta_str = f"  ETA ~{m}m{s:02d}s"

        print(f"\n[{i:>3}/{n_total}] {ticker}{eta_str}")

        try:
            df_1m    = fetch_bars(ticker, 1, "minute", start, end)
            df_daily = fetch_bars(ticker, 1, "day",    start, end)
        except Exception as e:
            print(f"    ⚠ fetch error: {e} — skipping")
            skipped.append(ticker)
            continue

        if df_1m.empty or df_daily.empty:
            print(f"    ⚠ no data — skipping")
            skipped.append(ticker)
            continue

        try:
            raw_trades, df_trades = run_ticker(ticker, df_1m, df_daily, CONFIG)
        except Exception as e:
            print(f"    ⚠ engine error: {e} — skipping")
            skipped.append(ticker)
            continue

        n_exits = 0
        wr = avg_r = 0.0
        if df_trades is not None and not df_trades.empty:
            exits  = df_trades[df_trades["Type"] == "Exit long"]
            n_exits = len(exits)
            wins   = (exits["Net P&L %"] > 0).sum()
            wr     = wins / n_exits * 100 if n_exits else 0
            avg_r  = exits["Net P&L %"].mean() if n_exits else 0

        elapsed = time.time() - t0
        times.append(elapsed)
        tag = "✓" if n_exits else "–"
        print(f"    {tag} {n_exits} signals  WR {wr:.1f}%  avg {avg_r:+.3f}%  [{elapsed:.1f}s]")

        all_trades[ticker] = df_trades

        # Rate-limit courtesy between tickers (only when we had to fetch)
        was_cached = all(
            _cache_path(ticker, m, ts, start, end).exists()
            for m, ts in [(1, "minute"), (1, "day")]
        )
        if not was_cached:
            time.sleep(TICKER_SLEEP)

    # ── Aggregate summary
    total_elapsed = time.time() - total_t0
    print("\n" + "=" * 65)
    print(f"  Done. {len(all_trades)} tickers processed in {total_elapsed:.0f}s")
    if skipped:
        print(f"  Skipped ({len(skipped)}): {', '.join(skipped)}")

    valid = {t: df for t, df in all_trades.items()
             if df is not None and not df.empty}
    if valid:
        all_exits = pd.concat(
            [df[df["Type"] == "Exit long"] for df in valid.values()],
            ignore_index=True,
        )
        n     = len(all_exits)
        wins  = (all_exits["Net P&L %"] > 0).sum()
        wr    = wins / n * 100 if n else 0
        avg_r = all_exits["Net P&L %"].mean() if n else 0
        print(f"\n  ── NDX100 Aggregate ──")
        print(f"  Total signals: {n:,}")
        print(f"  Win rate     : {wr:.1f}%")
        print(f"  Avg return   : {avg_r:+.3f}%")
        print(f"  Avg win      : {all_exits.loc[all_exits['Net P&L %'] > 0, 'Net P&L %'].mean():+.3f}%")
        print(f"  Avg loss     : {all_exits.loc[all_exits['Net P&L %'] <= 0, 'Net P&L %'].mean():+.3f}%")

    # ── Export
    output = Path(__file__).parent / CONFIG["output_file"]
    export_results(all_trades, str(output), chart_tf,
                   detail_sheets=CONFIG.get("detail_sheets", 25))

    print(f"\nTotal time : {total_elapsed:.0f}s ({total_elapsed/60:.1f} min)")
    print(f"Output     : {output}")


if __name__ == "__main__":
    main()
