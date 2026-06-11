"""
TPS Strategy — Improvement Ideas Analysis from Existing Backtest Data
======================================================================
What this can and cannot measure:

IDEA 1 (TP2 +4 ATR):
  Since the strategy uses process_orders_on_close=true, TradingView calculates
  favorable excursion on closing prices. For a winning TP2 trade, the max close
  during the trade equals the TP2 trigger price — so favorable_excursion ≈ TP2 return
  for winning trades. This means we CAN'T use favorable excursion to estimate whether
  price would have reached +4 ATR. HOWEVER, we CAN check for overshoot: when a TP2
  limit order fills, the closing bar may be above the TP2 price (gap/continuation).
  We measure this "overshoot" across all winning TP2 trades to estimate potential gains.

The main value of this script:
  - Reports baseline TP1 vs TP2 leg performance split
  - Quantifies TP2 overshoot (favorable excursion - TP2 return on winning trades)
  - Shows near-miss profile on losing TP2 trades (how close they came to positive)
  - All of this informs WHICH improvements are likely highest leverage

Note: For ideas 2,3,4,5,8 — a TradingView re-backtest is required. Pine Script
variants are in the TPS Project folder ready to paste and run.
"""

import openpyxl

TICKERS = ["AAPL", "GOOGL", "NVDA", "TSLA", "AMZN", "MSFT", "AMD", "ORCL", "NFLX"]

FILES = {
    "78m": "TPS Score Strategy Backtesting V2 5.12.26.xlsx",
    "195m": "TPS Score Strategy Backtesting V2 195M 5.18.26.xlsx",
}

def find_col(sheet, header):
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col).value
        if val and str(val).strip() == header:
            return col
    return None

def analyze(sheet, ticker):
    col_sig = find_col(sheet, "Signal")
    col_type = find_col(sheet, "Type")
    col_n = find_col(sheet, "Net P&L %")
    col_p = find_col(sheet, "Favorable excursion %")
    col_q = find_col(sheet, "Adverse excursion %")

    if not all([col_sig, col_n, col_p]):
        return None

    tp1_returns, tp2_win_returns, tp2_loss_returns = [], [], []
    tp2_overshoot = []          # favorable_excursion - net_pnl for winning TP2 trades
    tp2_loss_fav_exc = []       # favorable excursion on losing TP2 trades

    for row in range(2, sheet.max_row + 1):
        sig = sheet.cell(row=row, column=col_sig).value
        typ = sheet.cell(row=row, column=col_type).value if col_type else None
        if typ and "entry" in str(typ).lower():
            continue

        raw_n = sheet.cell(row=row, column=col_n).value
        raw_p = sheet.cell(row=row, column=col_p).value
        if raw_n is None:
            continue

        n = float(raw_n)
        p = float(raw_p) if raw_p is not None else 0.0

        if sig == "TP1":
            tp1_returns.append(n)
        elif sig == "TP2":
            if n > 0:
                tp2_win_returns.append(n)
                tp2_overshoot.append(p - n)
            else:
                tp2_loss_returns.append(n)
                tp2_loss_fav_exc.append(p)

    def avg(lst): return sum(lst)/len(lst) if lst else 0.0
    def pct(a, b): return 100*a/b if b else 0.0

    n_tp1 = len(tp1_returns)
    n_tp2w = len(tp2_win_returns)
    n_tp2l = len(tp2_loss_returns)
    n_tp2 = n_tp2w + n_tp2l

    return {
        "ticker": ticker,
        # TP1 leg
        "tp1_n": n_tp1,
        "tp1_wr": pct(sum(1 for r in tp1_returns if r > 0), n_tp1),
        "tp1_avg": avg(tp1_returns),
        # TP2 leg
        "tp2_n": n_tp2,
        "tp2_wr": pct(n_tp2w, n_tp2),
        "tp2_avg_win": avg(tp2_win_returns),
        "tp2_avg_loss": avg(tp2_loss_returns),
        "tp2_avg_all": avg(tp2_win_returns + tp2_loss_returns),
        # Idea 1 signal: overshoot on winning TP2 trades
        "overshoot_avg": avg(tp2_overshoot),       # how much past TP2 price closed above target
        "overshoot_pct_pos": pct(sum(1 for x in tp2_overshoot if x > 0.01), n_tp2w),
        # Near-miss profile on losing TP2 trades
        "loss_fav_avg": avg(tp2_loss_fav_exc),     # avg favorable move before stopping
        "loss_near_miss_pct": pct(sum(1 for x in tp2_loss_fav_exc if x > 0.5), n_tp2l),  # >0.5% favorable
    }


def run(label, filepath):
    print(f"\n{'='*80}")
    print(f"  {label} — TP1 vs TP2 LEG BREAKDOWN")
    print(f"{'='*80}")
    print(f"\n{'Ticker':<7} {'TP1 Trades':>10} {'TP1 WR%':>8} {'TP1 Avg%':>9} | "
          f"{'TP2 Trades':>10} {'TP2 WR%':>8} {'TP2 Avg%':>9} {'TP2W Avg%':>10} {'TP2L Avg%':>10}")
    print("-" * 90)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    results = []
    for t in TICKERS:
        if t not in wb.sheetnames:
            continue
        r = analyze(wb[t], t)
        if r:
            results.append(r)
            print(f"{r['ticker']:<7} {r['tp1_n']:>10} {r['tp1_wr']:>8.1f}% {r['tp1_avg']:>9.3f}% | "
                  f"{r['tp2_n']:>10} {r['tp2_wr']:>8.1f}% {r['tp2_avg_all']:>9.3f}% "
                  f"{r['tp2_avg_win']:>10.3f}% {r['tp2_avg_loss']:>10.3f}%")
    wb.close()

    if results:
        def avg_r(k): return sum(r[k] for r in results) / len(results)
        print("-" * 90)
        print(f"{'ALL':<7} {int(avg_r('tp1_n')):>10} {avg_r('tp1_wr'):>8.1f}% {avg_r('tp1_avg'):>9.3f}% | "
              f"{int(avg_r('tp2_n')):>10} {avg_r('tp2_wr'):>8.1f}% {avg_r('tp2_avg_all'):>9.3f}% "
              f"{avg_r('tp2_avg_win'):>10.3f}% {avg_r('tp2_avg_loss'):>10.3f}%")

    print(f"\n\n{'='*80}")
    print(f"  {label} — IDEA 1 SIGNAL: TP2 OVERSHOOT + NEAR-MISS ANALYSIS")
    print(f"{'='*80}")
    print(f"  Overshoot = favorable_excursion - TP2_return on winning trades")
    print(f"  (If >0: price closed above TP2 target on the exit bar → wider TP2 captures more)")
    print(f"  Near-miss = losing TP2 trades where favorable move was ≥ 0.5%")
    print()
    print(f"{'Ticker':<7} {'TP2 Wins':>9} {'Overshoot Avg%':>15} {'Overshoot>0 %':>14} | "
          f"{'TP2 Losses':>11} {'Fav Exc Avg%':>13} {'NearMiss>0.5%':>14}")
    print("-" * 85)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    results2 = []
    for t in TICKERS:
        if t not in wb.sheetnames:
            continue
        r = analyze(wb[t], t)
        if r:
            results2.append(r)
            print(f"{r['ticker']:<7} {r['tp2_n'] - int(r['tp2_n']*(1-r['tp2_wr']/100)):>9} "
                  f"{r['overshoot_avg']:>15.4f}% {r['overshoot_pct_pos']:>13.1f}% | "
                  f"{int(r['tp2_n']*(1-r['tp2_wr']/100)):>11} {r['loss_fav_avg']:>13.3f}% "
                  f"{r['loss_near_miss_pct']:>13.1f}%")
    wb.close()


if __name__ == "__main__":
    import os
    os.chdir("/Users/stephenprior/Documents/GitHub/brain/Projects/TPS Project")
    for label, fp in FILES.items():
        run(label, fp)

    print(f"\n\n{'='*80}")
    print("  INTERPRETATION & IMPLICATIONS FOR EACH IMPROVEMENT IDEA")
    print(f"{'='*80}")
    print("""
IDEA 1 — TP2 +4 ATR (wider take profit):
  Key signal: "Overshoot Avg%" above. If this is near 0 on most tickers, closing prices
  rarely exceed the TP2 target, meaning +4 ATR requires more favorable continuation.
  Verdict: Requires TradingView re-backtest. High risk of reducing TP2 hit rate.
  Alternative: Consider adding a trailing stop instead (Idea 2) which harvests more
  of winning trades WITHOUT sacrificing TP2 hit rate.

IDEA 2 — Trailing stop after TP1:
  Signal from data: Look at "TP2 Loss Avg%" — these are trades where TP1 was hit but
  TP2 was stopped at adjusted stop (+1 ATR). A trailing stop would have protected more
  of these trades. Requires TradingView re-backtest.
  Verdict: Likely positive — see Pine Script variant.

IDEA 3 — Score threshold 70 (vs 65):
  No score data in spreadsheet → cannot estimate. Must re-run in TradingView.
  Hypothesis: Fewer trades, higher win rate. May improve risk-adjusted returns.
  Verdict: Requires re-backtest. See Pine Script variant.

IDEA 4 — Squeeze weight rebalancing (78m ↑, 15m ↓):
  Changes which trades fire → cannot estimate from existing data.
  Hypothesis: More weight on chart-TF squeeze → signal closer to the actual entry TF.
  Verdict: Requires re-backtest. See Pine Script variant.

IDEA 5 — Volume filter (volume > 20-bar SMA):
  Changes which trades fire → cannot estimate from existing data.
  Hypothesis: Fewer false breakouts, modest trade reduction, cleaner entries.
  Verdict: Requires re-backtest. See Pine Script variant.

IDEA 8 — Squeeze duration 5–12 bars (vs 3–20):
  Changes which trades fire → cannot estimate from existing data.
  Hypothesis: Tighter sweet spot for squeeze duration → reduces noise at extremes.
  Verdict: Requires re-backtest. See Pine Script variant.

READY-TO-RUN PINE SCRIPTS (paste into TradingView one at a time):
  TPS_v2_idea1_tp2_4atr.pine
  TPS_v2_idea2_trailing_stop.pine
  TPS_v2_idea3_score70.pine
  TPS_v2_idea4_sqz_weights.pine
  TPS_v2_idea5_volume.pine
  TPS_v2_idea8_sqz_duration.pine
""")
